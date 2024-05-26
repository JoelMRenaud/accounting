from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import sqlite3

workbook = Workbook()
ws = workbook.active

con = sqlite3.connect("sql.db")
cur = con.cursor()

res = cur.execute("SELECT * from company_info")
data = res.fetchall()

res = cur.execute("SELECT COUNT(*) FROM accounts")
account_num = int(''.join(map(str, res.fetchall()[0])))

res = cur.execute("SELECT * FROM accounts ORDER BY account_number")
accounts = res.fetchall()

ws['A1'] = data[0][0]
curCell = ws['A1']
curCell.alignment = Alignment(horizontal='center')

ws['A2'] = 'Trial Balance'
curCell = ws['A2']
curCell.alignment = Alignment(horizontal='center')

ws['A3'] = 'January 31, 20--'
curCell = ws['A3']
curCell.alignment = Alignment(horizontal='center')

ws['A4'] = 'Account'
curCell = ws['A4']
curCell.alignment = Alignment(horizontal='center')

ws['E4'] = 'Acc. No.'
curCell = ws['E4']
curCell.alignment = Alignment(horizontal='center')

ws['F4'] = 'Debit'
curCell = ws['F4']
curCell.alignment = Alignment(horizontal='center')

ws['G4'] = 'Credit'
curCell = ws['G4']
curCell.alignment = Alignment(horizontal='center')

for x in range(3):
    ws.merge_cells(start_row=x+1, start_column=1, end_row=x+1, end_column=7)

for x in range(50):
    ws.merge_cells(start_row=x+1, start_column=1, end_row=x+1, end_column=4)

sumLeft = 0
sumRight = 0
for x in range(account_num):
    if int(accounts[x][3]) != 0:
        ws['A' + str(x+5)] = accounts[x][1]
        ws['E' + str(x+5)] = accounts[x][0]
        if int(accounts[x][3]) > 0:
            ws['F' + str(x+5)] = abs(int(accounts[x][3]))
            sumLeft = sumLeft + int(accounts[x][3])
        if int(accounts[x][3]) < 0:
            ws['G' + str(x+5)] = abs(int(accounts[x][3]))
            sumRight = sumLeft + int(accounts[x][3])

ws['F' + str(account_num + 5)] = sumLeft
ws['G' + str(account_num + 5)] = sumLeft

# Define a bold font and double border
bold_font = Font(bold=True)
double_border = Border(top=Side(style='thick'), bottom=Side(style='double'))

# Apply the font and border to the total row
ws['F' + str(account_num + 5)].font = bold_font
ws['G' + str(account_num + 5)].font = bold_font

ws['F' + str(account_num + 5)].border = double_border
ws['G' + str(account_num + 5)].border = double_border

workbook.save(filename="trial.xlsx")
con.close()